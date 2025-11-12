# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from typing import cast

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'


#workbook = Workbook() # criando um novo arquivo
workbook: Workbook = load_workbook(WORKBOOK_PATH) # carregando o arquivo existente


sheetname = 'Estudantes' # Nome da planilha

# criando uma nova planilha
worksheet: Worksheet = cast(Worksheet, workbook.create_sheet(title=sheetname))

#selecionando a planilha ativa (a primeira criada por padrão)
worksheet: Worksheet = cast(Worksheet, workbook.active)

# Renomeando a planilha ativa
#worksheet.title = sheetname

# removendo a planilha padrão criada
#workbook.remove(workbook['Sheet'])

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

#loop para adicionar os dados dos estudantes ( forma manual e mais complexa)
# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

# Adicionando os dados dos estudantes (forma simples)
#for student in students:
#    worksheet.append(student)



# Lendo os dados da planilha
for row in worksheet.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3):
    for cell in row:
        print(cell.value, end='\t ')
    print()  # Nova linha após cada linha de dados

#worksheet.append(['Carlos', 14, 7.5]) # Adicionando mais um estudante
#worksheet['A6'] = 'Carlos'  # Adicionando mais um estudante
#worksheet['B6'] = 14 # Idade
#worksheet['C6'] = 7.5 # Nota

#worksheet['b2'] = 15 # Alterando a idade do João




# Salvando o arquivo
#workbook.save(WORKBOOK_PATH)   