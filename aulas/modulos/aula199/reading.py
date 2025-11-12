from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()

sheetname = 'Estudantes' # Nome da planilha

# criando uma nova planilha
worksheet: Worksheet = cast(Worksheet, workbook.create_sheet(title=sheetname))

#selecionando a planilha ativa (a primeira criada por padrão)
worksheet: Worksheet = cast(Worksheet, workbook.active)

# Renomeando a planilha ativa
#worksheet.title = sheetname

# removendo a planilha padrão criada
#workbook.remove(workbook['Sheet'])

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)